
# Sourav Bhatti
#6/1/2022



from tkinter import N, RIDGE, S, Label, Listbox, filedialog
from tkinter.filedialog import askopenfile, askopenfilename
import pandas as pd
import datetime
import ctypes
from tkinter import CENTER, END, Button, Checkbutton, Frame, IntVar, Menu, StringVar, font, simpledialog, ttk, messagebox
from tkinter.ttk import Combobox, Treeview
import requests
import json
import tkinter as tk
import pyperclip
import openpyxl



root = tk.Tk()
root.geometry('1200x900+50+50')
root.title("Data Retriever V3.0.1 by Sourav Bhatti")
root.resizable(0,0)
ctypes.windll.shcore.SetProcessDpiAwareness(1)
root.configure(bg="grey16")

style = ttk.Style()
style.theme_use('clam')

text = ("Times new roman", 18, 'bold')
comboBoxFont =('Courier New', '15', "bold")
comboOption = StringVar()


qrCode =  StringVar()
tk.Label(root, text= "" ,  width = 15,bg="grey16" ).grid( column= 0, row = 0)
tk.Label(root, text= "Please select the Data you want to check." ,font=text,  width = 35, bg="grey16",fg='white' ).grid( column= 2, row = 0, pady=10)


combolist = Combobox(root, textvariable=comboOption, font=text, width=35, height=15, state='readonly' )
combolist.grid(column=2, row=1, pady=10)
combolist['values']=["MacId Duplicate Check","MacId","Firmware Version","PCBA", "Ship Date", "Battery Type", "Sim Card Carrier","Sim Card Number","IMEI Gateway", "Region", "Team Viewer ID"]
combolist.current(0)
root.option_add("*TCombobox*Listbox*Font", comboBoxFont)   # Change the font of the list

tk.Label(root, text= "Please Enter the QR code below." ,font= text, bg="grey16", fg='white').grid( column= 2, row = 2,pady=10)
qrEntry = tk.Entry(root, textvariable= qrCode , font=comboBoxFont, width = 40, bd=5, relief='ridge')
qrEntry.grid( column= 2, row = 5,pady=10)
tk.Label(root, text= "" ,  width = 40,height=3, bg="grey16").grid( column= 2, row = 8)


lstbxlabel = Label(root,text="Rejected Milestones", font= ("Times new roman", 14, 'bold'))
lstbxlabel.grid(column=3 , row= 8, padx=20, sticky=S )
lstbx = Listbox(root, height=20, width=30, border=10, selectmode=tk.EXTENDED)
lstbx.grid(column=3 , row= 9, padx=20)

lstBxVsb = ttk.Scrollbar(root, orient="vertical", command=lstbx.yview)
lstBxVsb.place(x=1140, y=310, height=320+20)
lstbx.configure(yscrollcommand=lstBxVsb.set)


tk.Label(root, text= "Total Number of Rejects" ,font= ("Times new roman", 13, 'bold'), width = 15 ).grid( column= 3, row = 10, ipadx=12, padx=10)
countLabel = tk.Label(root, text= "0" ,  width = 10,bg="red", font=('Courier New', '13', "bold"))
countLabel.grid( column= 3, row = 11, pady= 10)

batchButton = Button(root, text="Batch data",font= ("Times new roman", 13, 'bold'),  width=20, command= lambda: batchMacId()) 
batchButton.grid(column=3, row=1)


clearButton = Button(root, text="Clear",font= ("Times new roman", 13, 'bold'), relief="groove",bg= "yellow",  width=10, command= lambda: cleartree()) 
clearButton.grid(column=2, row=12)

tk.Label(root, text= "Good to Ship." ,font= ("Times new roman", 13, 'bold'), width = 15 ).grid( column= 2, row = 10, ipadx=12, padx=10, pady=10)
shipcountLabel = tk.Label(root, text= "0" ,bg="green",  width = 10, font=('Courier New', '13', "bold"))
shipcountLabel.grid( column= 2, row = 11, sticky=N)

#To count the rejects.
lstBoxCount = lstbx.index("end")
countLabel.config(text=f"{lstBoxCount}")
 

#####################################
#           TreeView                #
#####################################


# Calling pack method w.r.to vertical
# scrollbar

 
# Configuring treeview

style.configure("mystyle.Treeview", highlightthickness=50, bd=30, font=('Times new roman', 13, 'bold')) # Modify the font of the body
style.configure("mystyle.Treeview.Heading", font=('Times new roman', 15,'bold')) # Modify the font of the headings

displaydata = Treeview(root,style = 'mystyle.Treeview', columns=['QR CODE', 'Data Type','Data'], show='headings',
                        height=20, selectmode='extended', takefocus="Any")
displaydata.grid(column=2, row=9)
displaydata.column("# 1", anchor=CENTER, width=300)
displaydata.heading("# 1", text="QR CODE")
displaydata.column("# 2", anchor=CENTER, width=200)
displaydata.heading("# 2", text="Data Type")
displaydata.column("# 3", anchor=CENTER, width=300)
displaydata.heading("# 3", text="Data")

vsb = ttk.Scrollbar(root, orient="vertical", command=displaydata.yview)
vsb.place(x=900, y=300, height=378+20)
displaydata.configure(yscrollcommand=vsb.set)

listLast2Mac = []

def main():
    if comboOption.get() == "MacId Duplicate Check":
        macIdDuplicateCheck()
    else:
        getMacId() 

def macIdDuplicateCheck():

    
    data = pd.read_csv("MilestoneLog.csv", usecols=["QR code"], index_col=False)    #  read csv  sheet_name='Sheet1',
    df = pd.DataFrame(data)
    qrCodeList = df["QR code"].values.tolist()

    if qrCode.get() in qrCodeList:
            messagebox.showerror("Already Scanned!", "!!!!!Already Scanned!!!!!")
        
    
    else:

        url = "http://vmprdate.eastus.cloudapp.azure.com:9000/api/v1/manifest/?qrcode=" + qrCode.get()
        r = requests.get(url)
        rawjson = json.loads(r.text)
        if rawjson['error'] == False:
            data = rawjson['data']

                                
            macid = data[0]['macid']      
            last2Mac = macid[-2:]
            if last2Mac in listLast2Mac:
                messagebox.showerror("Last Mac Id Matched", f"Last 2 Mac Id Digits Matched.\n\nDo not ship Milestone with QR code {qrCode.get()}")
                lstbx.insert(tk.END, qrCode.get())
                lstbx.yview_moveto(1)
                status = "Duplicate"
                
                
            else:
                listLast2Mac.append(last2Mac)
                displaydata.insert('', END,
                        values=[qrCode.get(),comboOption.get(), macid])
                displaydata.yview_moveto(1)
                status = "Good to ship"
                

            tstamp = datetime.datetime.today().strftime("%m/%d/%y")
        

            # wb = openpyxl.load_workbook("MilestoneLog.xlsx")
            # sheet= wb.active
            # maxRow = sheet.max_row

            data = {"Date":[f"{tstamp}"],"QR code":[f"{qrCode.get()}"],"Data Type":[f"{comboOption.get()}"], "Data Value":[f'{macid}'], "Status":[f'{status}'] }
            df = pd.DataFrame(data)

            # with pd.ExcelWriter("MilestoneLog.xlsx", engine= "openpyxl", if_sheet_exists='overlay', mode="a") as writer:
            df.to_csv("MilestoneLog.csv", mode="a", header=None, index=False)
            # df.to_csv("MilestoneLog.csv", startrow=maxRow, startcol=0, header=None, index=False)

            qrEntry.delete(0,END)
    
    qrEntry.delete(0,END)
    # return
    lstBoxCount = lstbx.index("end")
    countLabel.config(text=f"{lstBoxCount}")
    shipcountLabel.config(text=f"{len(displaydata.get_children())}")

    # countLabel.config(text=f"{len(displaydata.get_children())}")


def getMacId():

    url = "http://vmprdate.eastus.cloudapp.azure.com:9000/api/v1/manifest/?qrcode=" + qrCode.get()
    r = requests.get(url)
    rawjson = json.loads(r.text)
    
    
    if rawjson['error'] == False:
        data = rawjson['data']


        # if comboOption.get() == "MacId Duplicate Check" :
                              
        #     macid = data[0]['macid']      
        #     last2Mac = macid[-2:]
        #     if last2Mac in listLast2Mac:
        #         messagebox.showerror("Last Mac Id Matched", f"Last 2 Mac Id Digits Matched.\n\nDo not ship Milestone with QR code {qrCode.get()}")
        #         lstbx.insert(tk.END, qrCode.get())
        #         status = "Duplicate"
                
        #     else:
        #         listLast2Mac.append(last2Mac)
        #         status = "Good to ship"
            

            
        if comboOption.get() == "MacId" :
            macid = data[0]['macid']      # Very Important Line
            status = ""
    
        elif comboOption.get() =="PCBA" :
            macid = data[0]['PCBA'] 
            status = ""

        elif comboOption.get() =="Firmware Version" :
            macid = data[0]['fw_version'] 
            status = ""

        elif comboOption.get() =="IMEI Gateway" :
            macid = data[0]['IMEInumber'] 
            status = ""

        elif comboOption.get() =="Battery Type" :
            macid = data[0]['batteryType'] 
            status = ""

        elif comboOption.get() =="Ship Date" :
            macid = data[0]['shipDate'][0:10] 
            status = ""

        elif comboOption.get() =="Sim Card Carrier" :
            macid = data[0]['comment'] 
            string = str(macid)
            listOfString = string.split("|")
            macid = listOfString[1]
            status = ""
        elif comboOption.get() =="Team Viewer ID" :
            macid = data[0]['comment'] 
            string = str(macid)
            listOfString = string.split("|")
            macid = listOfString[2]
            status = ""
        elif comboOption.get() =="Region" :
            macid = data[0]['comment'] 
            string = str(macid)
            listOfString = string.split("|")
            macid = listOfString[9]
            status = ""
        elif comboOption.get() =="Sim Card Number" :
            macid = data[0]['SIMcard'] 
            status = ""
        
                
        if macid != "None" and macid != "":
            # macids.append(macid)
            
            displaydata.insert('', END,
                    values=[qrCode.get(),comboOption.get(), macid])
            displaydata.yview_moveto(1)                             #Credit Goes to Mark....
            
        else:
            messagebox.showerror("Macid not found for ",qrCode.get())
        serverResponse = json.dumps(data)

        tstamp = datetime.datetime.today().strftime("%m/%d/%y")

        # wb = openpyxl.load_workbook("Log.xlsx")
        # sheet= wb.active
        # maxRow = sheet.max_row

        data = {"Date":[f"{tstamp}"],"QR code":[f"{qrCode.get()}"],"Data Type":[f"{comboOption.get()}"], "Data Value":[f'{macid}'], "Status":[f'{status}'] }
        df = pd.DataFrame(data)

        # with pd.ExcelWriter("Log.xlsx", engine= "openpyxl", if_sheet_exists='overlay', mode="a") as writer:
        df.to_csv("Log.csv",mode="a", header=None, index=False)
        # with open("Log.xlsx", "ab") as file:
        #     data = {"Date":[f"{tstamp}"],"QR code":[f"{qrCode.get()}"],"Data Type":[f"{comboOption.get()}"], "Data Value":[f'{macid}'], "Status":[f'{status}'] }
        #     df = pd.DataFrame(data)
        #     df.to_excel(file, index=False, header=False )
    shipcountLabel.config(text=f"{len(displaydata.get_children())}")
    qrEntry.delete(0,END)



    
    
root.bind('<Return>', lambda event: main())

def cleartree():

    messagebox.askokcancel("Clear the Data on Display", "Are you sure to clear window?")
    for item in displaydata.get_children():
        displaydata.delete(item)


# def copytree():

#     root.clipboard_clear()
#     if displaydata.focus():
#         curItem = displaydata.focus()
#         root.clipboard_append(f" (QR CODE) {displaydata.item(curItem)['values'][0]} | (MACID) - {displaydata.item(curItem)['values'][1]}\n")
#     else:
#         for items in displaydata.get_children():
#             #current_selection = displaydata.focus()
#             root.clipboard_append(f" (QR CODE) {displaydata.item(items)['values'][0]} | (MACID) - {displaydata.item(items)['values'][1]}\n" )
        
        
#     messagebox.showinfo("Please paste before closing program.", root.clipboard_get())


# def readexcelfile():
    

#     excelfile = filedialog.askopenfilename()
#     df = pd.read_excel(excelfile)
    
#     for value in df:

#         url = "http://vmprdate.eastus.cloudapp.azure.com:9000/api/v1/manifest/?qrcode=" + value
#         r = requests.get(url)
#         rawjson = json.loads(r.text)

        

#         if rawjson['error'] == False:
#             data = rawjson['data']


def batchMacId():
    try: 

        excelfile = filedialog.askopenfilename(filetypes=[("csv files","*.csv")])
        df = pd.read_csv(excelfile, usecols=["QR code"], index_col=False)
        qrCodeList = df["QR code"].values.tolist()    

        for i in range(len(qrCodeList)):

            # url = "http://vmprdate.eastus.cloudapp.azure.com:9000/api/v1/manifest/?qrcode=" + value
            # r = request.get(url)
            # rawjson = json.loads(r.text)

            qrcode = str(qrCodeList[i])
        
            url = "http://vmprdate.eastus.cloudapp.azure.com:9000/api/v1/manifest/?qrcode=" + qrcode
            r = requests.get(url)
            rawjson = json.loads(r.text)
            
            # data = rawjson["data"]
            # macid = data[0]["macid"]
            # print(f"{qrcode} | {macid}")
            if rawjson['error'] == False:
                data = rawjson['data']

                if comboOption.get() == "MacId Duplicate Check" :
                    messagebox.showerror("Feature not Available", " Batch Processing is not available for MacId Duplicate Check!")
                    return
            
                if comboOption.get() == "MacId" :
                    macid = data[0]['macid']      # Very Important Line
                    status = ""
            
                elif comboOption.get() =="PCBA" :
                    macid = data[0]['PCBA'] 
                    status = ""

                elif comboOption.get() =="Firmware Version" :
                    macid = data[0]['fw_version'] 
                    status = ""

                elif comboOption.get() =="IMEI Gateway" :
                    macid = data[0]['IMEInumber'] 
                    status = ""

                elif comboOption.get() =="Battery Type" :
                    macid = data[0]['batteryType'] 
                    status = ""

                elif comboOption.get() =="Ship Date" :
                    macid = data[0]['shipDate'] 
                    status = ""

                elif comboOption.get() =="Sim Card Carrier" :
                    macid = data[0]['comment'] 
                    string = str(macid)
                    listOfString = string.split("|")
                    macid = listOfString[1]
                    status = ""
                elif comboOption.get() =="Team Viewer ID" :
                    macid = data[0]['comment'] 
                    string = str(macid)
                    listOfString = string.split("|")
                    macid = listOfString[2]
                    status = ""
                elif comboOption.get() =="Region" :
                    macid = data[0]['comment'] 
                    string = str(macid)
                    listOfString = string.split("|")
                    macid = listOfString[9]
                    status = ""
                elif comboOption.get() =="Sim Card Number" :
                    macid = data[0]['SIMcard'] 
                    status = ""
                
                        
                if macid != "None" and macid != "":
                    # macids.append(macid)
                    
                    displaydata.insert('', END,
                            values=[qrcode,comboOption.get(), macid])
                    displaydata.yview_moveto(1)
                    
                else:
                    messagebox.showerror("Macid not found for ",qrCode.get())
                serverResponse = json.dumps(data)

                tstamp = datetime.datetime.today().strftime("%m/%d/%y")

                # wb = openpyxl.load_workbook("Log.xlsx")
                # sheet= wb.active
                # maxRow = sheet.max_row

                data = {"Date":[f"{tstamp}"],"QR code":[f"{qrcode}"],"Data Type":[f"{comboOption.get()}"], "Data Value":[f'{macid}'], "Status":[f'{status}'] }
                df = pd.DataFrame(data)

                # with pd.ExcelWriter("Log.xlsx", engine= "openpyxl", if_sheet_exists='overlay', mode="a") as writer:
                df.to_csv("Log.csv",mode="a", header=None, index=False)
    except FileNotFoundError:

        return                       





#################################################################
    # copiedString = ''
    # for line in displaydata.get_children():
        
    #     for value in displaydata.item(line)['values']:
    #         copiedString += f"{value}, "

    # pyperclip.copy(copiedString)

    # messagebox.showinfo("Copied", pyperclip.paste())
#################################################################

# tk.Label(root, text= "" ,  width = 15,bg="grey16" ).grid( column= 2, row = 10)
# tk.Label(root, text= "" ,  width = 15,bg="grey16" ).grid( column= 2, row = 11)

# Button(root, text='Copy',width=15, command= lambda :copytree()).grid(column=2, row= 12)

root.mainloop()


